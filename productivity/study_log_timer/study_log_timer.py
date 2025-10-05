# Gitバージョン管理のテスト
import tkinter as tk
from tkinter import simpledialog, messagebox
import datetime as dt
import openpyxl as xl
import os

# -----------------------------
# グローバル変数・ファイル設定
# -----------------------------
root = tk.Tk()
root.title("学習記録アプリ")
root.geometry("500x300")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
filename = os.path.join(BASE_DIR, "study_log.xlsx")

start_time = None
study_time = dt.timedelta()   #時間の長さを表す関数, 一時停止したときまでの累積時間
is_paused = False
record_start_time = None
subject_options = ["指定なし"]  # デフォルト値

# -----------------------------
# タイマー関連関数
# -----------------------------
def time_counter():
    """カウントアップタイマー"""
    global is_paused, study_time, start_time
    if start_time is None:
        return
    if not is_paused:
        elpased = dt.datetime.now() - start_time + study_time  # 学習時間 = 現在時刻 - スタートボタンを押した時刻 + 総学習時間
        timer_label.config(text=str(elpased).split('.')[0])
        root.after(1000, time_counter)  # 1秒後にtime_counter()を呼び出す

def press_start_button():
    """スタートボタン処理"""
    global start_time
    subject_frame.pack_forget()
    start_time = dt.datetime.now()
    start_button.pack_forget()  # スタートボタンを隠す
    break_button.pack(side="left",padx=20)
    end_button.pack(side="left",padx=20)
    timer_frame.pack(expand=True)
    label.config(text="""[休憩]=タイマーを一時停止します\n
    [終了]=終了し学習内容を記録します""")
    time_counter()
    subject_label.forget()

def press_break_button():
    """休憩ボタン処理"""
    global is_paused, study_time, start_time
    is_paused = True
    study_time += dt.datetime.now() - start_time
    break_button.config(text="再開", bg="#81C784", command=press_restart_button)

def press_restart_button():
    """再開ボタン処理"""
    global is_paused,  start_time
    if start_time is None:
        return # スタートしていなければ何もしない
    is_paused = False
    start_time = dt.datetime.now()  # 再開時の現在時刻を取得
    break_button.config(text="休憩", bg="#4FC3F7", command=press_break_button)
    time_counter()

def reset_timer():
    """タイマーをリセット"""
    global start_time, study_time, is_paused
    start_time = None
    study_time = dt.timedelta()
    is_paused = False
    timer_label.config(text="0:00:00")

# -----------------------------
# Excel関連関数
# -----------------------------
def save_to_excel():
    """学習内容をExcelに記録"""
    global start_time, study_time, is_paused, record_start_time, input_win
    content = input_text.get("1.0", "end-1c")   # 1.0→テキストの先頭(1行目0文字目)、end-1c→最後の改行を除いた最後の文字まで
    if not content.strip(): # 前後の空白や改行を削除
        return  # 空文字だったら何もせず終了
    
    # Excelを開く or 新規作成
    if os.path.exists(filename):
        wb = xl.load_workbook(filename)
    else:
        wb = xl.Workbook()

    # recordシート作成
    if "record" not in wb.sheetnames:
        ws = wb.create_sheet("record")
        ws.append(["日付", "開始時刻", "終了時刻", "学習時間（分）", "学習科目", "学習内容"])
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
    else:
        ws = wb["record"]
        
    # 2行目に新規記録
    ws.insert_rows(2)
    ws["A2"] = dt.datetime.now().strftime("%Y/%m/%d")
    ws["B2"] = record_start_time.strftime("%H:%M")
    ws["C2"] = dt.datetime.now().strftime("%H:%M")
    ws["D2"] = int(study_time.total_seconds() // 60)
    ws["E2"] = subject_var.get()
    ws["F2"] = content
    wb.save(filename)

    # UIリセット
    input_win.destroy()
    break_button.forget()
    end_button.forget()
    start_button.pack()
    timer_frame.forget()
    subject_frame.pack(expand=True)
    label.config(text="学習科目を選択してください")

def press_end_button():
    global record_start_time
    record_start_time = start_time  # reset前に開始時刻を保存
    record()  # 入力ウィンドウを開く
    reset_timer()  # タイマーをリセット

def record():
    global input_win, input_text
    input_win = tk.Toplevel()
    # input_win.geometry("500x100")
    input_label = tk.Label(input_win, text="改行は使用できません")
    input_label.pack(side="top")
    input_frame = tk.Frame(input_win)
    input_frame.pack(expand=True)
    input_text = tk.Text(input_frame, width=50, height=5)
    input_text.pack(pady=5, fill="x")   # 横幅いっぱいに広げる
    record_button = tk.Button(input_frame, text="記録", font=("Arial", 18), command=save_to_excel)
    record_button.pack(side="bottom",padx=20)

# -----------------------------
# 科目リスト操作
# -----------------------------
def load_subjects_from_excel():
    """Excelから科目リストを読込"""
    global subject_options
    if not os.path.exists(filename):
        return  # まだファイルがなければ何もしない
    try:
        wb = xl.load_workbook(filename)
        if "subject" in wb.sheetnames:
            ws = wb["subject"]
            subject_options = [cell.value for cell in ws["A"] if cell.value]
            if not subject_options:
                subject_options = ["指定なし"]
    except Exception as e:
        print("科目リスト読込失敗:", e)

def save_subjects_to_excel():
    """科目リストをExcelに保存"""
    wb = None
    if os.path.exists(filename):
        wb = xl.load_workbook(filename)
    else:
        wb = xl.Workbook()

    if "subject" not in wb.sheetnames:
        ws = wb.create_sheet("subject")
    else:
        ws = wb["subject"]
        for row in ws["A1:A100"]:
            for cell in row:
                cell.value = None
    for i, s in enumerate(subject_options, start=1):
        ws[f"A{i}"] = s
    wb.save(filename)

def update_subject_menu():
    """OptionMenuを更新"""
    menu = subject_menu["menu"]     # OptionMenuの「現在値を取得」
    menu.delete(0, "end")   # 現在のメニューを削除
    for option in subject_options:   # リストをループ
        menu.add_command(
            label=option,
            command=lambda value=option: subject_var.set(value)
        )
    
def add_subject():
    """科目追加"""
    new_subject = simpledialog.askstring("", "追加する科目を入力してください")
    # ifに変数そのものを書くと真偽値として評価される、[None, "", 0, [], {}]の場合Flase=リストに存在しない場合
    # もし(new_subjectがFlase)かつ(new_subjectがsubject_optionsにない)場合
    if new_subject and new_subject not in subject_options:
        subject_options.append(new_subject) #リストに追加
        update_subject_menu()    # OptionMenuを更新
        subject_var.set(new_subject)    # 追加した科目を選択状態にする
        save_subjects_to_excel()    # 追加後に即保存
    elif new_subject in subject_options:
        messagebox.showinfo("", "既にリストに存在します")

def delete_subject():
    """科目削除"""
    to_delete = subject_var.get()   # 現在選択中の科目を取得
    if to_delete in subject_options:
        subject_options.remove(to_delete)   # リストから削除
        update_subject_menu()
        if subject_options:
            subject_var.set(subject_options[0]) # 最初の科目を選択
        else:
            subject_var.set("")
        save_subjects_to_excel()    # 削除後に足保存

def open_excel():
    """学習記録(Excel)を開く"""
    if os.path.exists(filename):
        os.startfile(filename)
    else:
        messagebox.showarning("", "記録ファイルが存在しません")

# -----------------------------
# UI作成
# -----------------------------
label = tk.Label(root, text="学習科目を選択してください")
label.pack(side="top", pady=10)

subject_frame = tk.Frame(root)
subject_frame.pack(expand=True)
subject_menu_frame = tk.Frame(subject_frame)
subject_menu_frame.pack(side="top", pady=5)
# subject_label = tk.Label(subject_frame)
# subject_label.pack(side="left", padx=5)

load_subjects_from_excel()

subject_var = tk.StringVar(value=subject_options[0])    # tk.StringVar→文字列専用の変数、value→初期値
subject_menu = tk.OptionMenu(subject_menu_frame, subject_var, *subject_options)
subject_menu.config(font=("Arial", 24))
subject_menu.pack(side="top", pady=5)
# subject_options = ["指定なし"]

add_button = tk.Button(subject_frame, text="追加", font=("Arial", 14), command=add_subject)
add_button.pack(side="left", padx=5)

delete_button = tk.Button(subject_frame, text="削除", font=("Arial", 14), command=delete_subject)
delete_button.pack(side="left", padx=5)

open_excel_button = tk.Button(subject_frame, text="記録ファイルを開く", font=("Arial", 14), command=open_excel)
open_excel_button.pack(side="left", padx=5)

timer_frame = tk.Frame(root)
timer_frame.pack(expand=True)   # 縦方向に余白を使って中央に寄せる
timer_label = tk.Label(timer_frame, text="00:00:00", font=("Arial", 72))
timer_label.pack()
timer_frame.pack_forget()   # 非表示にしておく

button_frame = tk.Frame(root)
button_frame.pack(side="bottom", pady=10)
start_button = tk.Button(button_frame, text="スタート", font=("Arial", 18), command=press_start_button)
break_button = tk.Button(button_frame, text="休憩", font=("Arial", 18), bg="#4FC3F7", command=press_break_button)
end_button = tk.Button(button_frame, text="終了", font=("Arial", 18), command=press_end_button)
start_button.pack(side="left",padx=10)

root.mainloop()
